# -*- coding = utf-8 -*-
import json
import os
import sqlite3
import time
from datetime import datetime
import shutil
import openpyxl
import requests
import win32api
import win32con
import threading
from openpyxl.styles import Font, PatternFill

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/92.0.4515.107 Safari/537.36 Edg/92.0.902.62",
    "Content-Type": "application/json",
    "Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",
}
abspath = os.path.dirname(os.path.abspath(__file__))
downList = []


def writexls(bids):
    flag = 0
    if not os.path.exists(abspath + '\\Res\\信息.xlsx'):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '采购公告'
        wb.save(abspath + '\\Res\\信息.xlsx')
    ws_open = openpyxl.load_workbook(abspath + '\\Res\\信息.xlsx')
    for one_bid in bids:
        bid = [i for i in one_bid[0:9] if i is not None]
        sheet_name = one_bid[-1]
        pattern = one_bid[-2]
        if sheet_name not in ws_open.sheetnames:
            ws_open.create_sheet(sheet_name)
        ws = ws_open[sheet_name]
        if sheet_name == "采购公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '采购项目名称'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '采购项目编号'
            ws.column_dimensions['C'].width = 20
            ws['D1'] = '公告发布时间'
            ws.column_dimensions['D'].width = 13
            ws['E1'] = '采购文件获取截止时间'
            ws.column_dimensions['E'].width = 23
            ws['F1'] = '开启应答文件时间'
            ws.column_dimensions['F'].width = 23
            ws['G1'] = '网页链接'
            ws.column_dimensions['G'].width = 90
            ws['H1'] = '采购单位'
            ws.column_dimensions['H'].width = 30
        elif sheet_name == "招标公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '招标项目名称'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '招标项目编号'
            ws.column_dimensions['C'].width = 20
            ws['D1'] = '公告发布时间'
            ws.column_dimensions['D'].width = 13
            ws['E1'] = '招标文件获取截止时间'
            ws.column_dimensions['E'].width = 23
            ws['F1'] = '开启应答文件时间'
            ws.column_dimensions['F'].width = 23
            ws['G1'] = '公告文件'
            ws.column_dimensions['G'].width = 22
            ws['H1'] = '网页链接'
            ws.column_dimensions['H'].width = 90
            ws['I1'] = '采购单位'
            ws.column_dimensions['I'].width = 30
        elif sheet_name == "中标（成交）结果公告" or "推荐中标候选人公示" or "资格预审公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '项目名'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '发布时间'
            ws.column_dimensions['C'].width = 21
            ws['D1'] = '网页链接'
            ws.column_dimensions['D'].width = 90
            ws['E1'] = '采购单位'
            ws.column_dimensions['E'].width = 30
        if len(bid) > 2:
            flag = flag + 1
            ws.append(bid)
            if pattern == '1':
                color_fill = PatternFill(fill_type='solid', fgColor="AACF91")
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=ws.max_row, column=j).fill = color_fill
            if "库存" in bid[1]:
                font = Font(bold=True, color="FF0000")
                ws.cell(row=ws.max_row, column=2).font = font
    if flag:
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\twrite to excel!")
        ws_open.save(abspath + '\\Res\\信息.xlsx')
        shutil.copy(abspath + '\\Res\\信息.xlsx ', abspath + '\\国网招采信息.xlsx')
        return True
    else:
        return False


def DB2Excel(num):
    conn = sqlite3.connect(abspath + '\\Res\\db.db')
    towrite = []
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        # 2.执行SQL操作
        cursor.execute("select * from (select * from hisdata order by id desc limit 0,?)as tbl order by id ", (num,))
        data = cursor.fetchall()
        if data is not None:
            for i in data:
                read_db = list(i[1:13])
                towrite.append(read_db)
            return towrite
        else:
            return 0
    except sqlite3.DatabaseError as error:
        # 4. 回滚数据库事物
        print(error)
        conn.rollback()
    finally:
        # 5. 关闭数据连接
        conn.close()


def getDownloadUrl(lstid):
    download_url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/downLoadBid?noticeId=' + str(
        lstid) + '&noticeDetId='
    BinPath = os.path.join(abspath, 'Res')
    os.chdir(BinPath)
    if not os.path.exists(abspath + '\\Download\\' + str(lstid) + '.zip'):
        cmd = 'wget.exe --content-disposition "%s" -O "%s"' % (
            download_url, abspath + "\\Download\\" + str(lstid) + ".zip")
        os.system(cmd)


# noinspection PyBroadException
def getidtype():
    url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/noteList'
    getlist = []
    requestdata = [
        '{"index":1,"size":40,"firstPageMenuId":"2018032900295987","purOrgStatus":"","purOrgCode":"","purType":"",'
        '"orgId":"","key":"","orgName":""}',
        # 采购公告
        '{"index":1,"size":40,"firstPageMenuId":"2018032700291334","purOrgStatus":"","purOrgCode":"","purType":"",'
        '"orgId":"","key":"","orgName":""}',
        # 招标公告
        '{"index":1,"size":40,"firstPageMenuId":"2018032700290425","purOrgStatus":"","purOrgCode":"","purType":"",'
        '"orgId":"","key":"","orgName":""}',
        # 资格预审公告
        '{"index":1,"size":40,"firstPageMenuId":"2018060501171107","orgId":"","key":"","year":"","orgName":""}',
        # 推荐中标候选人公示
        '{"index":1,"size":40,"firstPageMenuId":"2018060501171111","orgId":"","key":"","year":"","orgName":""}']
    # 中标（成交）结果公告
    for reqdata in requestdata:
        while True:
            try:
                res = requests.post(url=url, data=reqdata, timeout=20)
                break
            except requests.exceptions.ConnectionError:
                print('ConnectionError -- please wait 3 seconds')
                time.sleep(3)
            except requests.exceptions.ChunkedEncodingError:
                print('ChunkedEncodingError -- please wait 3 seconds')
                time.sleep(3)
            except:
                print('Unfortunitely -- An Unknow Error Happened, Please wait 3 seconds')
                time.sleep(3)
        result = json.loads(res.text)
        for i in range(39, -1, -1):
            x = result['resultValue']["noteList"][i]['id']
            j = result['resultValue']["noteList"][i]['doctype']
            mId = result['resultValue']["noteList"][i]["firstPageMenuId"]
            getlist.append({
                'noticid': x,
                'doctype': j,
                'menuID': mId,
            })
    return getlist


'''
采购公告                   "2018032900295987"
招标公告                   "2018032700291334"
资格预审公告               "2018032700290425"
中标（成交）结果公告       "2018060501171111"
推荐中标候选人公示         "2018060501171107"
'''


def parseURL(lstid, doctype, menuid):
    Url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/'
    bidlist = []
    key = ['gettime', 'noticeId', 'TYPE_NAME']
    typename = ""
    if menuid == 2018032900295987:
        typename = "采购公告"
    elif menuid == 2018032700291334:
        typename = "招标公告"
    elif menuid == 2018032700290425:
        typename = "资格预审公告"
    elif menuid == 2018060501171111:
        typename = "中标（成交）结果公告"
    elif menuid == 2018060501171107:
        typename = "推荐中标候选人公示"
    if doctype == "doci-bid":
        getNoticeBidurl = Url + 'getNoticeBid'
        res = json.loads(requests.post(url=getNoticeBidurl, data=json.dumps(lstid), headers=headers).text)
        PurType = res["resultValue"]["notice"]["PUR_TYPE_NAME"]
        if PurType == "物资":
            bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            bidlist.append(res["resultValue"]["notice"]["PURPRJ_NAME"])
            bidlist.append(res["resultValue"]["notice"]["PUB_TIME"])
            zipname = str(res["resultValue"]["notice"]["ONLINE_BID_NOTICE_ID"]) + ".zip"
            if menuid == 2018032900295987:
                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                       'BIDBOOK_BUY_END_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                bidlist.append(res["resultValue"]["notice"]["PURPRJ_CODE"])
                bidlist.append(res["resultValue"]["notice"]["BIDBOOK_BUY_END_TIME"])
                bidlist.append(res["resultValue"]["notice"]["BIDBOOK_SELL_BEGIN_TIME"])
            elif menuid == 2018032700291334:
                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                       'BIDBOOK_BUY_END_TIME', 'zip', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                bidlist.append(res["resultValue"]["notice"]["PURPRJ_CODE"])
                bidlist.append(res["resultValue"]["notice"]["BIDBOOK_BUY_END_TIME"])
                bidlist.append(res["resultValue"]["notice"]["BIDBOOK_SELL_BEGIN_TIME"])
                downList.append(str(lstid))
                bidlist.append('=HYPERLINK("' + abspath + '\\Download\\' + zipname + '","' + zipname + '")')
            elif menuid == 2018032700290425:
                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
            bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-bid/' + str(lstid) + '_' + str(menuid))
            bidlist.append('0')
            bidlist.append(res["resultValue"]["notice"]["BID_ORG"])
            print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid))
        else:
            key = ['gettime', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
            bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            bidlist.append('0')
            bidlist.append(res["resultValue"]["notice"]["BID_ORG"])
    elif doctype == "doci-change":
        getChangeBidurl = Url + 'getChangeBid'
        res = json.loads(requests.post(url=getChangeBidurl, data=json.dumps(lstid), headers=headers).text)
        PurType = res["resultValue"]["origNotice"]["PUR_TYPE_NAME"]
        bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if PurType == "物资":
            bidlist.append(res["resultValue"]["origNotice"]["PURPRJ_NAME"])
            bidlist.append(res["resultValue"]["origNotice"]["PUB_TIME"])
            if res["resultValue"]["chgNotice"] is not None:
                if menuid == 2018032900295987:
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                           'BIDBOOK_BUY_END_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                    bidlist.append(res["resultValue"]["chgNotice"]["PURPRJ_CODE"])
                    bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_BUY_END_TIME"])
                    bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                elif menuid == 2018032700290425:
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                elif menuid == 2018032700291334:
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                           'BIDBOOK_BUY_END_TIME', 'zip', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                    bidlist.append(res["resultValue"]["chgNotice"]["PURPRJ_CODE"])
                    bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_BUY_END_TIME"])
                    bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                    zipname = str(lstid) + ".zip"
                    downList.append(str(lstid))
                    bidlist.append('=HYPERLINK("' + abspath + '\\Download\\' + zipname + '","' + zipname + '")')
                print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename + "变更".ljust(11) + "\t" + str(
                    lstid))
            else:
                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'zip', 'URL', 'histype', 'ORGNAME', 'noticeId',
                       'TYPE_NAME']
                bidlist.append("waring")
            bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-change/' + str(lstid) + '_' + str(menuid))
        else:
            key = ['gettime', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
        bidlist.append('1')
        bidlist.append(res["resultValue"]["origNotice"]["BID_ORG"])
    elif doctype == "doci-win":
        getNoticeWinurl = Url + 'getNoticeWin'
        key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
        res = json.loads(requests.post(url=getNoticeWinurl, data=json.dumps(lstid), headers=headers).text)
        bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        bidlist.append(res["resultValue"]["notice"]["TITLE"])
        bidlist.append(res["resultValue"]["notice"]["PUB_TIME"])
        bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-win/' + str(lstid) + '_' + str(menuid))
        bidlist.append('0')
        bidlist.append(res["resultValue"]["notice"]["ORG_NAME_SM4"])
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid))
    elif doctype == "doc-com":
        getDocUrl = Url + 'getDoc'
        key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
        res = json.loads(requests.post(url=getDocUrl, data=json.dumps(lstid), headers=headers).text)
        bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        bidlist.append(res["resultValue"]["doc"]["title"])
        bidlist.append(res["resultValue"]["doc"]["noticePublishTime"])
        bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doc-com/' + str(lstid) + '_' + str(menuid))
        bidlist.append('0')
        bidlist.append(res["resultValue"]["doc"]["publishOrgName"])
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid))
    elif doctype == "doc-spec":
        getDocurl = Url + 'getDoc'
        res = json.loads(requests.post(url=getDocurl, data=json.dumps(lstid), headers=headers).text)
        key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
        bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        bidlist.append(res["resultValue"]["doc"]["purOrgName"])
        bidlist.append(res["resultValue"]["doc"]["noticePublishTime"])
        bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doc-spec/' + str(lstid) + '_' + str(menuid))
        bidlist.append('0')
        bidlist.append(res["resultValue"]["doc"]["publishOrgName"])
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid))
    bidlist.append(str(lstid))
    bidlist.append(typename)
    insertdata = zip(key, bidlist)
    insert_data(dict(insertdata))


def insert_data(values):
    conn = sqlite3.connect(abspath + '\\Res\\db.db')
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        columns = ', '.join(values.keys())
        placeholders = ':' + ', :'.join(values.keys())
        query = 'INSERT INTO hisdata (%s) VALUES (%s)' % (columns, placeholders)
        cursor.execute(query, values)
        conn.commit()
    except sqlite3.DatabaseError as error:
        print(error)
        conn.rollback()
    finally:
        conn.close()


def isindb():
    conn = sqlite3.connect(abspath + '\\Res\\db.db')
    relist = []
    try:
        cursor = conn.cursor()
        results = cursor.execute("select noticeId from hisdata")
        data = results.fetchall()
        for i in data:
            relist.append(i[0])
        return relist
    except sqlite3.DatabaseError as error:
        print(error)
        conn.rollback()
    finally:
        conn.close()


def readid():
    conn = sqlite3.connect(abspath + '\\Res\\db.db')
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        # 2.执行SQL操作
        # noinspection SqlResolve
        cursor.execute("select id from hisdata order by id desc limit 0,1")
        data = cursor.fetchone()
        if data is not None:
            return data[0]
        else:
            return 0
    except sqlite3.DatabaseError as error:
        # 4. 回滚数据库事物
        print(error)
        conn.rollback()
    finally:
        # 5. 关闭数据连接
        conn.close()


# noinspection PyBroadException
def create_table():
    conn = sqlite3.connect(abspath + '\\Res\\db.db')
    try:
        create_tb_cmd = '''
        CREATE TABLE IF NOT EXISTS hisdata (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT
                                        NOT NULL,
        gettime                 TEXT,
        PURPRJ_NAME             TEXT,
        PURPRJ_CODE             TEXT,
        PUB_TIME                TEXT,
        BIDBOOK_SELL_BEGIN_TIME TEXT,
        BIDBOOK_BUY_END_TIME    TEXT,
        zip                     TEXT,
        URL                     TEXT,
        ORGNAME                 TEXT,
        noticeId                TEXT,
        histype                 TEXT,
        TYPE_NAME        TEXT);
        '''
        conn.execute(create_tb_cmd)
    except:
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + '\tCreate table failed')
        return False
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\tdb Connected!")
    conn.close()


def create_dir_not_exist(path):
    if not os.path.exists(path):
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + '\t' + path + ' not exitsts ,created!')
        os.mkdir(path)


if __name__ == "__main__":
    create_dir_not_exist(abspath + '\\Res')
    create_dir_not_exist(abspath + '\\Download')
    create_table()
    '''
    reqdata = [
        '{"index":1,"size":40,"firstPageMenuId":"2018032900295987","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":"","orgName":""}',
        # 采购公告
        '{"index":1,"size":40,"firstPageMenuId":"2018032700291334","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":"","orgName":""}',
        # 招标公告
        '{"index":1,"size":40,"firstPageMenuId":"2018032700290425","purOrgStatus":"","purOrgCode":"","purType":"","orgId":"","key":"","orgName":""}',
        # 资格预审公告
        '{"index":1,"size":40,"firstPageMenuId":"2018060501171111","orgId":"","key":"","year":"","orgName":""}']  
        # 中标（成交）结果公告
    '''
    thread_list = []
    downthread_list = []
    start = time.time()
    idlist = isindb()
    startid = readid()
    getlst = getidtype()
    for t in getlst:
        if str(t["noticid"]) not in idlist:
            m = threading.Thread(target=parseURL, args=(t["noticid"], t["doctype"], t["menuID"]))
            thread_list.append(m)
    for m in thread_list:
        m.start()
    for m in thread_list:
        m.join()
    for down in downList:
        n = threading.Thread(target=getDownloadUrl, args=(down,))
        downthread_list.append(n)
    for n in downthread_list:
        n.start()
    for n in downthread_list:
        n.join()
    addnum = readid() - startid
    if addnum > 0:
        lstw = DB2Excel(addnum)
        if len(lstw) > 0:
            if writexls(lstw):
                win32api.MessageBox(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "有新的信息了！", "提醒", win32con.MB_OK)
            else:
                print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t没有新的信息")
    print('[info]耗时：%s' % (time.time() - start))
